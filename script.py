import openpyxl
path=r"c:\Users\sk670\Downloads\D2C_Marketing_Calculator (1).xlsx"
wb=openpyxl.load_workbook(path,data_only=False)
for sheet in wb.sheetnames:
    ws=wb[sheet]
    print(f"Sheet: {sheet}")
    for row in ws.iter_rows(min_row=1,max_col=10,max_row=20):
        for cell in row:
            if isinstance(cell.value,str) and cell.value.startswith('='):
                print(cell.coordinate,cell.value)
